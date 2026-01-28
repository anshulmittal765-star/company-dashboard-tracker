#!/usr/bin/env python3
"""
Company Dashboard Tracker
Scrapes financial data for watchlist companies and creates Excel dashboard
"""

import os
import sys
import time
import base64
import json
from datetime import datetime
from collections import defaultdict

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIGURATION
# ============================================================================

SCREENER_USERNAME = os.getenv('SCREENER_USERNAME')
SCREENER_PASSWORD = os.getenv('SCREENER_PASSWORD')
GOOGLE_SHEET_ID = os.getenv('GOOGLE_SHEET_ID')
GOOGLE_CREDENTIALS_BASE64 = os.getenv('GOOGLE_CREDENTIALS_BASE64')

# Watchlist URLs
WATCHLISTS = {
    "My Stonks": os.getenv('MY_STONKS_WATCHLIST_URL', ''),
    "Core Watchlist": os.getenv('CORE_WATCHLIST_URL', ''),
}

# ============================================================================
# SELENIUM SETUP
# ============================================================================

def setup_selenium():
    """Configure Chrome WebDriver"""
    print("🔧 Setting up Selenium...")
    options = Options()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('--window-size=1920,1080')
    return webdriver.Chrome(options=options)

def login_to_screener(driver):
    """Login to Screener.in"""
    print("🔐 Logging into Screener.in...")
    try:
        driver.get('https://www.screener.in/login/')
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.NAME, 'username'))
        )
        driver.find_element(By.NAME, 'username').send_keys(SCREENER_USERNAME)
        driver.find_element(By.NAME, 'password').send_keys(SCREENER_PASSWORD)
        driver.find_element(By.CSS_SELECTOR, 'button[type="submit"]').click()
        time.sleep(3)
        if "login" in driver.current_url.lower():
            print("❌ Login failed!")
            return False
        print("✅ Login successful!")
        return True
    except Exception as e:
        print(f"❌ Login error: {e}")
        return False

# ============================================================================
# SCRAPE WATCHLISTS
# ============================================================================

def scrape_watchlist_companies(driver, watchlist_url):
    """Get list of companies from a watchlist"""
    if not watchlist_url:
        return []
    
    try:
        print(f"📋 Scraping watchlist: {watchlist_url}")
        driver.get(watchlist_url)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "table"))
        )
        
        companies = []
        rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
        
        for row in rows:
            try:
                name_cell = row.find_element(By.CSS_SELECTOR, "td a")
                company_name = name_cell.text.strip()
                company_url = name_cell.get_attribute('href')
                
                if company_name and company_url:
                    companies.append({
                        'name': company_name,
                        'url': company_url
                    })
            except:
                continue
        
        print(f"✅ Found {len(companies)} companies")
        return companies
        
    except Exception as e:
        print(f"⚠️  Error scraping watchlist: {e}")
        return []

def get_all_watchlist_companies(driver):
    """Get all companies from all watchlists"""
    print("\n📊 Collecting companies from watchlists...")
    all_companies = []
    seen = set()
    
    for name, url in WATCHLISTS.items():
        if url:
            companies = scrape_watchlist_companies(driver, url)
            for company in companies:
                if company['name'] not in seen:
                    seen.add(company['name'])
                    all_companies.append(company)
    
    print(f"\n✅ Total unique companies: {len(all_companies)}")
    return all_companies

# ============================================================================
# SCRAPE COMPANY DATA
# ============================================================================

def scrape_company_data(driver, company):
    """Scrape financial data for a single company"""
    print(f"\n📈 Scraping: {company['name']}")
    
    try:
        driver.get(company['url'])
        time.sleep(2)
        
        data = {
            'name': company['name'],
            'url': company['url'],
            'current_price': None,
            'market_cap': None,
            'pe_ratio': None,
            'sector': None,
            'revenue_data': {},
            'profit_data': {},
            'margin_data': {},
            'quarterly_data': [],
        }
        
        # Extract basic info from top section
        try:
            # Current Price
            price_elem = driver.find_element(By.CSS_SELECTOR, "#top-ratios li:nth-child(1) .number")
            data['current_price'] = price_elem.text.strip()
            
            # Market Cap
            mcap_elem = driver.find_element(By.CSS_SELECTOR, "#top-ratios li:nth-child(3) .number")
            data['market_cap'] = mcap_elem.text.strip()
            
            # P/E Ratio
            pe_elem = driver.find_element(By.CSS_SELECTOR, "#top-ratios li:nth-child(5) .number")
            data['pe_ratio'] = pe_elem.text.strip()
            
            # Sector
            sector_elem = driver.find_element(By.CSS_SELECTOR, ".sub a")
            data['sector'] = sector_elem.text.strip()
            
        except Exception as e:
            print(f"  ⚠️  Could not extract basic info: {e}")
        
        # Extract 10-year data from tables
        try:
            # Find Profit & Loss table
            tables = driver.find_elements(By.CSS_SELECTOR, "section.card table")
            
            for table in tables:
                try:
                    # Check table header to identify which table it is
                    header = table.find_element(By.CSS_SELECTOR, "thead th").text.strip()
                    
                    rows = table.find_elements(By.CSS_SELECTOR, "tbody tr")
                    
                    for row in rows:
                        try:
                            cells = row.find_elements(By.TAG_NAME, "td")
                            if len(cells) < 2:
                                continue
                            
                            metric_name = cells[0].text.strip()
                            
                            # Extract yearly values
                            values = []
                            for cell in cells[1:]:
                                values.append(cell.text.strip())
                            
                            # Store based on metric name
                            if 'Sales' in metric_name or 'Revenue' in metric_name:
                                data['revenue_data'][metric_name] = values
                            elif 'Operating Profit' in metric_name or 'EBITDA' in metric_name:
                                data['profit_data'][metric_name] = values
                            elif 'Net Profit' in metric_name:
                                data['profit_data'][metric_name] = values
                            elif 'margin' in metric_name.lower():
                                data['margin_data'][metric_name] = values
                                
                        except:
                            continue
                            
                except:
                    continue
                    
        except Exception as e:
            print(f"  ⚠️  Could not extract table data: {e}")
        
        # Extract quarterly results
        try:
            quarterly_section = driver.find_element(By.CSS_SELECTOR, "section#quarters")
            q_table = quarterly_section.find_element(By.CSS_SELECTOR, "table")
            q_rows = q_table.find_elements(By.CSS_SELECTOR, "tbody tr")
            
            for q_row in q_rows[:8]:  # Last 8 quarters
                try:
                    cells = q_row.find_elements(By.TAG_NAME, "td")
                    if len(cells) >= 3:
                        quarter = {
                            'period': cells[0].text.strip(),
                            'sales': cells[1].text.strip(),
                            'profit': cells[2].text.strip()
                        }
                        data['quarterly_data'].append(quarter)
                except:
                    continue
                    
        except Exception as e:
            print(f"  ⚠️  Could not extract quarterly data: {e}")
        
        print(f"  ✅ Data extracted for {company['name']}")
        return data
        
    except Exception as e:
        print(f"  ❌ Error scraping {company['name']}: {e}")
        return None

def scrape_all_companies(driver, companies):
    """Scrape data for all companies"""
    print("\n📊 Scraping financial data for all companies...")
    
    all_data = []
    success_count = 0
    
    for i, company in enumerate(companies, 1):
        print(f"\n[{i}/{len(companies)}]")
        data = scrape_company_data(driver, company)
        
        if data:
            all_data.append(data)
            success_count += 1
        
        # Rate limiting
        time.sleep(1)
    
    print(f"\n✅ Successfully scraped {success_count}/{len(companies)} companies")
    return all_data

# ============================================================================
# GOOGLE SHEETS
# ============================================================================

def get_google_credentials():
    """Get Google credentials"""
    try:
        creds_json = base64.b64decode(GOOGLE_CREDENTIALS_BASE64).decode('utf-8')
        creds_dict = json.loads(creds_json)
        return Credentials.from_service_account_info(
            creds_dict,
            scopes=['https://www.googleapis.com/auth/spreadsheets']
        )
    except Exception as e:
        print(f"❌ Error loading credentials: {e}")
        raise

def save_to_google_sheets(all_data):
    """Save raw data to Google Sheets"""
    print("\n📊 Saving to Google Sheets...")
    
    try:
        creds = get_google_credentials()
        service = build('sheets', 'v4', credentials=creds)
        
        # Prepare data
        headers = ['Company', 'Sector', 'Current Price', 'Market Cap', 'P/E Ratio', 'URL']
        rows = [headers]
        
        for data in all_data:
            rows.append([
                data['name'],
                data.get('sector', ''),
                data.get('current_price', ''),
                data.get('market_cap', ''),
                data.get('pe_ratio', ''),
                data['url']
            ])
        
        # Clear and update
        service.spreadsheets().values().clear(
            spreadsheetId=GOOGLE_SHEET_ID,
            range='Sheet1!A1:Z1000'
        ).execute()
        
        result = service.spreadsheets().values().update(
            spreadsheetId=GOOGLE_SHEET_ID,
            range='Sheet1!A1',
            valueInputOption='RAW',
            body={'values': rows}
        ).execute()
        
        print(f"✅ Saved {len(rows)} rows to Google Sheets")
        return True
        
    except Exception as e:
        print(f"❌ Error saving to Google Sheets: {e}")
        return False

# ============================================================================
# EXCEL GENERATION
# ============================================================================

def create_excel_dashboard(all_data):
    """Create interactive Excel dashboard"""
    print("\n📊 Creating Excel dashboard...")
    
    try:
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        dashboard_sheet = wb.create_sheet("Dashboard", 0)
        data_sheet = wb.create_sheet("Raw Data", 1)
        company_list_sheet = wb.create_sheet("Company List", 2)
        
        # ========================================
        # RAW DATA SHEET
        # ========================================
        print("  📄 Creating Raw Data sheet...")
        
        # Headers
        headers = ['Company', 'Sector', 'Price', 'Market Cap', 'P/E', 'URL']
        for col, header in enumerate(headers, 1):
            cell = data_sheet.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Data
        for row, data in enumerate(all_data, 2):
            data_sheet.cell(row, 1, data['name'])
            data_sheet.cell(row, 2, data.get('sector', ''))
            data_sheet.cell(row, 3, data.get('current_price', ''))
            data_sheet.cell(row, 4, data.get('market_cap', ''))
            data_sheet.cell(row, 5, data.get('pe_ratio', ''))
            data_sheet.cell(row, 6, data['url'])
        
        # ========================================
        # COMPANY LIST SHEET (for dropdown)
        # ========================================
        print("  📄 Creating Company List...")
        
        company_list_sheet.cell(1, 1, "Company Name")
        for row, data in enumerate(all_data, 2):
            company_list_sheet.cell(row, 1, data['name'])
        
        # ========================================
        # DASHBOARD SHEET
        # ========================================
        print("  📄 Creating Dashboard sheet...")
        
        # Title
        dashboard_sheet.merge_cells('A1:F1')
        title_cell = dashboard_sheet.cell(1, 1, "COMPANY FINANCIAL DASHBOARD")
        title_cell.font = Font(size=18, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        dashboard_sheet.row_dimensions[1].height = 30
        
        # Dropdown label
        dashboard_sheet.cell(3, 1, "Select Company:")
        dashboard_sheet.cell(3, 1).font = Font(bold=True, size=12)
        
        # Dropdown cell (B3)
        dropdown_cell = dashboard_sheet.cell(3, 2)
        dropdown_cell.value = all_data[0]['name'] if all_data else ""
        
        # Add data validation for dropdown
        from openpyxl.worksheet.datavalidation import DataValidation
        
        dv = DataValidation(
            type="list",
            formula1=f"='Company List'!$A$2:$A${len(all_data)+1}",
            allow_blank=False
        )
        dv.add(dropdown_cell)
        dashboard_sheet.add_data_validation(dv)
        
        # Key Metrics Section
        dashboard_sheet.cell(5, 1, "Key Metrics")
        dashboard_sheet.cell(5, 1).font = Font(bold=True, size=14)
        
        metrics = [
            ("Current Price:", "=VLOOKUP($B$3,'Raw Data'!$A:$F,3,FALSE)"),
            ("Market Cap:", "=VLOOKUP($B$3,'Raw Data'!$A:$F,4,FALSE)"),
            ("P/E Ratio:", "=VLOOKUP($B$3,'Raw Data'!$A:$F,5,FALSE)"),
            ("Sector:", "=VLOOKUP($B$3,'Raw Data'!$A:$F,2,FALSE)"),
        ]
        
        for i, (label, formula) in enumerate(metrics, 6):
            dashboard_sheet.cell(i, 1, label)
            dashboard_sheet.cell(i, 1).font = Font(bold=True)
            dashboard_sheet.cell(i, 2, formula)
        
        # Instructions
        dashboard_sheet.cell(12, 1, "Instructions:")
        dashboard_sheet.cell(12, 1).font = Font(bold=True, size=12)
        dashboard_sheet.cell(13, 1, "1. Select a company from the dropdown above")
        dashboard_sheet.cell(14, 1, "2. All metrics will update automatically")
        dashboard_sheet.cell(15, 1, "3. Check 'Raw Data' sheet for all company information")
        
        # Column widths
        dashboard_sheet.column_dimensions['A'].width = 20
        dashboard_sheet.column_dimensions['B'].width = 30
        data_sheet.column_dimensions['A'].width = 25
        data_sheet.column_dimensions['B'].width = 20
        data_sheet.column_dimensions['F'].width = 50
        
        # Hide sheets
        company_list_sheet.sheet_state = 'hidden'
        
        # Save file
        filename = f"Company_Dashboard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        print(f"✅ Excel created: {filename}")
        return filename
        
    except Exception as e:
        print(f"❌ Error creating Excel: {e}")
        import traceback
        traceback.print_exc()
        return None

# ============================================================================
# MAIN
# ============================================================================

def main():
    """Main execution"""
    print("=" * 70)
    print("COMPANY DASHBOARD TRACKER")
    print("=" * 70)
    print(f"Started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    
    # Validate environment variables
    required_vars = {
        'SCREENER_USERNAME': SCREENER_USERNAME,
        'SCREENER_PASSWORD': SCREENER_PASSWORD,
        'GOOGLE_SHEET_ID': GOOGLE_SHEET_ID,
        'GOOGLE_CREDENTIALS_BASE64': GOOGLE_CREDENTIALS_BASE64
    }
    
    missing = [k for k, v in required_vars.items() if not v]
    if missing:
        print(f"❌ Missing environment variables: {', '.join(missing)}")
        sys.exit(1)
    
    driver = None
    
    try:
        # Setup
        driver = setup_selenium()
        
        # Login
        if not login_to_screener(driver):
            raise Exception("Login failed")
        
        # Get watchlist companies
        companies = get_all_watchlist_companies(driver)
        
        if not companies:
            print("⚠️  No companies found in watchlists!")
            sys.exit(0)
        
        # Scrape all companies
        all_data = scrape_all_companies(driver, companies)
        
        if not all_data:
            print("❌ No data scraped!")
            sys.exit(1)
        
        # Save to Google Sheets
        save_to_google_sheets(all_data)
        
        # Create Excel dashboard
        excel_file = create_excel_dashboard(all_data)
        
        # Summary
        print("\n" + "=" * 70)
        print("✅ SUMMARY")
        print("=" * 70)
        print(f"Companies scraped: {len(all_data)}")
        print(f"Excel file: {excel_file}")
        print(f"Google Sheet: https://docs.google.com/spreadsheets/d/{GOOGLE_SHEET_ID}")
        print(f"Completed at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("=" * 70)
        
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
        
    finally:
        if driver:
            driver.quit()
            print("\n🔒 Browser closed.")

if __name__ == "__main__":
    main()
